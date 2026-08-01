"""Geracao de relatorios e exportacoes (Excel e PDF).

Os dados sao montados uma unica vez em uma estrutura tabular neutra
(``cabecalhos`` + ``linhas``) e entregues a qualquer formato de saida. Assim
o mesmo relatorio sai em tela, Excel ou PDF sem duplicar a consulta nem a
regra de negocio.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from app.extensions import db
from app.models.enums import SituacaoCadastro, SituacaoMatricula
from app.models.estrutura import Serie, Turma
from app.models.matricula import Matricula
from app.models.pessoas import Aluno, Professor
from app.services import frequencia_service
from app.utils.validadores import formatar_cpf, formatar_telefone

# ---------------------------------------------------------------------------
# Estilos da planilha
# ---------------------------------------------------------------------------
FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=10)
FUNDO_CABECALHO = PatternFill("solid", fgColor="1A56DB")
BORDA_FINA = Border(*[Side(style="thin", color="D1D5DB")] * 4)
ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Dados dos relatorios
# ---------------------------------------------------------------------------
def relatorio_alunos(
    ano_letivo_id: int | None = None,
    turma_id: int | None = None,
    situacao: str | None = None,
) -> dict[str, Any]:
    """Listagem de alunos com turma, responsavel e contato."""
    consulta = db.session.query(Aluno).filter(Aluno.excluido_em.is_(None))

    if situacao:
        consulta = consulta.filter(Aluno.situacao == situacao)

    if turma_id or ano_letivo_id:
        sub = db.session.query(Matricula.aluno_id).filter(
            Matricula.situacao == SituacaoMatricula.ATIVA,
            Matricula.excluido_em.is_(None),
        )
        if turma_id:
            sub = sub.filter(Matricula.turma_id == turma_id)
        if ano_letivo_id:
            sub = sub.filter(Matricula.ano_letivo_id == ano_letivo_id)
        consulta = consulta.filter(Aluno.id.in_(sub))

    alunos = consulta.order_by(Aluno.nome_normalizado).all()

    linhas = []
    for aluno in alunos:
        turma = aluno.turma_atual
        responsavel = aluno.responsavel_principal
        linhas.append(
            [
                aluno.codigo,
                aluno.nome_exibicao,
                turma.identificacao_curta if turma else "Sem matricula",
                aluno.data_nascimento.strftime("%d/%m/%Y") if aluno.data_nascimento else "",
                str(aluno.idade) if aluno.idade is not None else "",
                formatar_cpf(aluno.cpf),
                formatar_telefone(aluno.celular or aluno.telefone),
                responsavel.nome_completo if responsavel else "",
                formatar_telefone(responsavel.celular or responsavel.telefone)
                if responsavel
                else "",
                aluno.situacao.rotulo,
            ]
        )

    return {
        "titulo": "Relatorio de Alunos",
        "cabecalhos": [
            "Codigo", "Nome", "Turma", "Nascimento", "Idade", "CPF",
            "Contato", "Responsavel", "Contato do responsavel", "Situacao",
        ],
        "linhas": linhas,
        "paisagem": True,
    }


def relatorio_turmas(ano_letivo_id: int | None = None) -> dict[str, Any]:
    """Panorama das turmas com ocupacao e disciplinas."""
    consulta = db.session.query(Turma).filter(
        Turma.excluido_em.is_(None), Turma.ativa.is_(True)
    )
    if ano_letivo_id:
        consulta = consulta.filter(Turma.ano_letivo_id == ano_letivo_id)

    turmas = consulta.join(Turma.serie).order_by(Serie.ordem, Turma.nome).all()

    linhas = []
    for turma in turmas:
        ocupados = turma.contar_matriculas_ativas()
        linhas.append(
            [
                turma.identificacao_curta,
                turma.serie.nome if turma.serie else "",
                turma.turno.rotulo,
                turma.sala.nome if turma.sala else "",
                turma.professor_regente.nome_exibicao if turma.professor_regente else "",
                str(ocupados),
                str(turma.capacidade),
                f"{turma.taxa_ocupacao:.1f}%".replace(".", ","),
                str(len([v for v in turma.turmas_disciplinas if v.ativa])),
            ]
        )

    return {
        "titulo": "Relatorio de Turmas",
        "cabecalhos": [
            "Turma", "Serie", "Turno", "Sala", "Regente",
            "Matriculados", "Capacidade", "Ocupacao", "Disciplinas",
        ],
        "linhas": linhas,
        "paisagem": True,
    }


def relatorio_professores() -> dict[str, Any]:
    """Corpo docente com formacao e carga horaria."""
    professores = (
        db.session.query(Professor)
        .filter(
            Professor.excluido_em.is_(None),
            Professor.situacao == SituacaoCadastro.ATIVO,
        )
        .order_by(Professor.nome_normalizado)
        .all()
    )

    linhas = []
    for professor in professores:
        disciplinas = ", ".join(d.nome for d in professor.disciplinas_lecionadas)
        linhas.append(
            [
                professor.registro_funcional,
                professor.nome_exibicao,
                professor.formacao or "",
                professor.titulacao or "",
                disciplinas or "Sem atribuicao",
                str(professor.carga_horaria_atribuida),
                str(professor.carga_horaria_semanal),
                formatar_telefone(professor.celular or professor.telefone),
                professor.email or "",
            ]
        )

    return {
        "titulo": "Relatorio de Professores",
        "cabecalhos": [
            "Registro", "Nome", "Formacao", "Titulacao", "Disciplinas",
            "Aulas atribuidas", "Carga contratual", "Contato", "E-mail",
        ],
        "linhas": linhas,
        "paisagem": True,
    }


def relatorio_matriculas(
    ano_letivo_id: int | None = None, situacao: str | None = None
) -> dict[str, Any]:
    """Movimentacao de matriculas do ano letivo."""
    consulta = db.session.query(Matricula).filter(Matricula.excluido_em.is_(None))

    if ano_letivo_id:
        consulta = consulta.filter(Matricula.ano_letivo_id == ano_letivo_id)
    if situacao:
        consulta = consulta.filter(Matricula.situacao == situacao)

    matriculas = (
        consulta.join(Aluno, Matricula.aluno_id == Aluno.id)
        .order_by(Aluno.nome_normalizado)
        .all()
    )

    linhas = [
        [
            matricula.numero,
            matricula.nome_aluno,
            matricula.turma.identificacao_curta if matricula.turma else "",
            matricula.data_matricula.strftime("%d/%m/%Y"),
            matricula.situacao.rotulo,
            matricula.resultado_final.rotulo,
            matricula.data_saida.strftime("%d/%m/%Y") if matricula.data_saida else "",
            matricula.escola_destino or matricula.escola_origem or "",
        ]
        for matricula in matriculas
    ]

    return {
        "titulo": "Relatorio de Matriculas",
        "cabecalhos": [
            "Numero", "Aluno", "Turma", "Data", "Situacao",
            "Resultado", "Saida", "Escola origem/destino",
        ],
        "linhas": linhas,
        "paisagem": True,
    }


def relatorio_frequencia(
    ano_letivo_id: int, frequencia_minima: float = 75.0
) -> dict[str, Any]:
    """Alunos com frequencia abaixo do minimo legal."""
    em_risco = frequencia_service.alunos_em_risco(ano_letivo_id, frequencia_minima)

    linhas = [
        [
            item["aluno"].codigo if item["aluno"] else "",
            item["aluno"].nome_exibicao if item["aluno"] else "",
            item["turma"].identificacao_curta if item["turma"] else "",
            str(item["total_aulas"]),
            str(item["total_faltas"]),
            f"{item['percentual']:.1f}%".replace(".", ","),
            item["matricula"].situacao.rotulo if item["matricula"] else "",
        ]
        for item in em_risco
    ]

    return {
        "titulo": "Relatorio de Frequencia - Alunos em Risco",
        "cabecalhos": [
            "Codigo", "Aluno", "Turma", "Aulas", "Faltas", "Frequencia", "Situacao",
        ],
        "linhas": linhas,
        "paisagem": False,
        "observacao": (
            f"Criterio: frequencia inferior a {frequencia_minima:.0f}% "
            "(minimo exigido pela LDB), considerando apenas alunos com pelo "
            "menos 10 aulas registradas."
        ),
    }


def relatorio_desempenho(ano_letivo_id: int) -> dict[str, Any]:
    """Media geral por turma, para acompanhamento pedagogico."""
    linhas_banco = (
        db.session.query(
            Turma.nome,
            Serie.nome,
            func.count(Matricula.id),
            func.avg(Matricula.media_geral),
            func.avg(Matricula.percentual_frequencia),
        )
        .select_from(Turma)
        .join(Serie, Turma.serie_id == Serie.id)
        .join(Matricula, Matricula.turma_id == Turma.id)
        .filter(
            Turma.ano_letivo_id == ano_letivo_id,
            Matricula.situacao == SituacaoMatricula.ATIVA,
            Matricula.excluido_em.is_(None),
        )
        .group_by(Turma.id, Turma.nome, Serie.nome, Serie.ordem)
        .order_by(Serie.ordem, Turma.nome)
        .all()
    )

    linhas = []
    for nome_turma, nome_serie, total, media, frequencia in linhas_banco:
        linhas.append(
            [
                f"{nome_serie} {nome_turma}",
                str(total),
                f"{float(media):.2f}".replace(".", ",") if media else "—",
                f"{float(frequencia):.1f}%".replace(".", ",") if frequencia else "—",
            ]
        )

    return {
        "titulo": "Relatorio de Desempenho por Turma",
        "cabecalhos": ["Turma", "Alunos", "Media geral", "Frequencia media"],
        "linhas": linhas,
        "paisagem": False,
        "observacao": (
            "Os valores refletem a ultima consolidacao de resultados. "
            "Consolide as turmas para atualizar os numeros."
        ),
    }


#: Catalogo de relatorios disponiveis na tela.
RELATORIOS: dict[str, dict[str, Any]] = {
    "alunos": {
        "nome": "Alunos",
        "descricao": "Listagem completa com turma, contato e responsavel.",
        "icone": "bi-people",
        "cor": "primaria",
        "permissao": "relatorio.administrativo",
    },
    "turmas": {
        "nome": "Turmas",
        "descricao": "Ocupacao, regente e disciplinas de cada turma.",
        "icone": "bi-diagram-3",
        "cor": "info",
        "permissao": "relatorio.administrativo",
    },
    "professores": {
        "nome": "Professores",
        "descricao": "Corpo docente, formacao e carga horaria atribuida.",
        "icone": "bi-person-video3",
        "cor": "sucesso",
        "permissao": "relatorio.administrativo",
    },
    "matriculas": {
        "nome": "Matriculas",
        "descricao": "Movimentacao de matriculas do ano letivo.",
        "icone": "bi-file-earmark-text",
        "cor": "alerta",
        "permissao": "relatorio.administrativo",
    },
    "frequencia": {
        "nome": "Frequencia em risco",
        "descricao": "Alunos abaixo da frequencia minima legal.",
        "icone": "bi-exclamation-triangle",
        "cor": "perigo",
        "permissao": "relatorio.academico",
    },
    "desempenho": {
        "nome": "Desempenho por turma",
        "descricao": "Media geral e frequencia media de cada turma.",
        "icone": "bi-graph-up",
        "cor": "primaria",
        "permissao": "relatorio.academico",
    },
}


def gerar_dados(chave: str, **filtros) -> dict[str, Any]:
    """Despacha para a funcao do relatorio solicitado."""
    from app.services.excecoes import RegistroNaoEncontrado

    ano_letivo_id = filtros.get("ano_letivo_id")

    if chave == "alunos":
        return relatorio_alunos(
            ano_letivo_id=ano_letivo_id,
            turma_id=filtros.get("turma_id"),
            situacao=filtros.get("situacao"),
        )
    if chave == "turmas":
        return relatorio_turmas(ano_letivo_id=ano_letivo_id)
    if chave == "professores":
        return relatorio_professores()
    if chave == "matriculas":
        return relatorio_matriculas(
            ano_letivo_id=ano_letivo_id, situacao=filtros.get("situacao")
        )
    if chave == "frequencia":
        if not ano_letivo_id:
            return {
                "titulo": "Relatorio de Frequencia",
                "cabecalhos": [],
                "linhas": [],
                "paisagem": False,
                "observacao": "Selecione um ano letivo para gerar este relatorio.",
            }
        return relatorio_frequencia(ano_letivo_id)
    if chave == "desempenho":
        if not ano_letivo_id:
            return {
                "titulo": "Relatorio de Desempenho",
                "cabecalhos": [],
                "linhas": [],
                "paisagem": False,
                "observacao": "Selecione um ano letivo para gerar este relatorio.",
            }
        return relatorio_desempenho(ano_letivo_id)

    raise RegistroNaoEncontrado("Relatorio nao encontrado.")


# ---------------------------------------------------------------------------
# Exportacao em Excel
# ---------------------------------------------------------------------------
def gerar_excel(dados: dict[str, Any]) -> BytesIO:
    """Converte a estrutura tabular em uma planilha formatada."""
    livro = Workbook()
    planilha = livro.active
    planilha.title = dados["titulo"][:31]  # limite do Excel

    # Titulo mesclado no topo.
    total_colunas = max(1, len(dados["cabecalhos"]))
    planilha.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=total_colunas
    )
    celula_titulo = planilha.cell(row=1, column=1, value=dados["titulo"])
    celula_titulo.font = Font(bold=True, size=13, color="1A56DB")
    celula_titulo.alignment = ALINHAMENTO_CENTRO

    planilha.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=total_colunas
    )
    celula_data = planilha.cell(
        row=2, column=1,
        value=f"Emitido em {date.today().strftime('%d/%m/%Y')} pelo SGE",
    )
    celula_data.font = Font(size=9, color="6B7280")
    celula_data.alignment = ALINHAMENTO_CENTRO

    # Cabecalho da tabela.
    linha_cabecalho = 4
    for indice, titulo in enumerate(dados["cabecalhos"], start=1):
        celula = planilha.cell(row=linha_cabecalho, column=indice, value=titulo)
        celula.font = FONTE_CABECALHO
        celula.fill = FUNDO_CABECALHO
        celula.alignment = ALINHAMENTO_CENTRO
        celula.border = BORDA_FINA

    # Dados.
    for numero_linha, linha in enumerate(dados["linhas"], start=linha_cabecalho + 1):
        for numero_coluna, valor in enumerate(linha, start=1):
            celula = planilha.cell(row=numero_linha, column=numero_coluna, value=valor)
            celula.border = BORDA_FINA
            celula.font = Font(size=9)

    # Largura das colunas conforme o conteudo, com teto para nao estourar.
    for indice, titulo in enumerate(dados["cabecalhos"], start=1):
        maior = len(str(titulo))
        for linha in dados["linhas"]:
            if indice <= len(linha):
                maior = max(maior, len(str(linha[indice - 1] or "")))
        planilha.column_dimensions[get_column_letter(indice)].width = min(maior + 3, 45)

    # Congela o cabecalho para facilitar a leitura de listas longas.
    planilha.freeze_panes = planilha.cell(row=linha_cabecalho + 1, column=1)

    if dados.get("observacao"):
        linha_obs = linha_cabecalho + len(dados["linhas"]) + 2
        celula = planilha.cell(row=linha_obs, column=1, value=dados["observacao"])
        celula.font = Font(size=8, italic=True, color="6B7280")

    buffer = BytesIO()
    livro.save(buffer)
    buffer.seek(0)
    return buffer


def responder_excel(buffer: BytesIO, nome_arquivo: str) -> Response:
    """Envia a planilha como download."""
    resposta = Response(
        buffer.getvalue(),
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resposta.headers["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    resposta.headers["Cache-Control"] = "private, no-store"
    return resposta
